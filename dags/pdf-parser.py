from __future__ import annotations
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
import pandas as pd
from pypdf import PdfReader 
import os
import re
import subprocess

from airflow.decorators import dag, task

####################################################################🕵️‍♂️Analyse de PDFs🕵️‍♂️####################################################################

TOOLS_PATH = os.environ.get("TOOLS_PATH")
WORKDIR_WORKFLOW = os.environ.get("WORKDIR_WORKFLOW")
OUTPUT_EXCEL = f"{WORKDIR_WORKFLOW}/Enquete-xxxxx-pdf.xlsx"

@dag(
    dag_id="pdf_investigation",
    start_date=datetime(2016,1,1),
    schedule=None,
    catchup=False,
)

def pdf_investigation():

    @task
    def list_pdfs(pdf_dir: str = WORKDIR_WORKFLOW) -> list[str]:
        pdfs = [os.path.join(pdf_dir, f) for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")]

        return pdfs
    
    @task
    def analyze_pdf(path: str) -> dict:
        filename = os.path.basename(path)

        reader = PdfReader(path)
        info = reader.metadata or {}

        def _clean_date(raw:str):
            if not raw or not raw.startswith("D:"):
                return None
            
            raw = raw[2:]

            m = re.match(r"(\d{14})([+-]\d{2})'?(\d{2})'?", raw)
            if not m:
                return raw
            
            dt_raw, tz_h, tz_m = m.groups()
            year = int(dt_raw[0:4])
            month = int(dt_raw[4:6])
            day = int(dt_raw[6:8])
            hour = int(dt_raw[8:10])
            minute = int(dt_raw[10:12])
            second = int(dt_raw[12:14])

            offset = timedelta(hours=int(tz_h), minutes=int(tz_m))
            tz = timezone(offset)
            dt = datetime(year, month, day, hour, minute, second, tzinfo=tz)
            return dt.isoformat()

        author = info.get("/Author")
        creator = info.get("/Creator")
        producer = info.get("/Producer")
        creation_date = _clean_date(info.get("/CreationDate"))
        mod_date = _clean_date(info.get("/ModDate"))

        with open(path,"rb") as f:
            data = f.read()

        try:
            text = data.decode("latin-1", errors="ignore")
        except Exception :
            text = data.decode("utf-8", errors="ignore")
        
        startxrefs = []
        for m in re.finditer(r"startxref\s+(\d+)", text):
            startxrefs.append(int(m.group(1)))
        startxrefs_count = len(startxrefs)

        trailers = []
        for m in re.finditer(r"<<[^>]*?/Type\s*/XRef.*?>>", text, flags=re.DOTALL):
            trailers.append(m.group(0))
        xref_blocks = []
        for m in re.finditer(r"<<\s*(?:.|\n)*?/Type\s*/XRef\s*(?:.|\n)*?>>", text):
            xref_blocks.append(m.group(0))

        trailers_count = len(trailers)

        revision_indices = []

        for block in xref_blocks:
            size_m = re.search(r"/Size\s+(\d+)", block)
            size = int(size_m.group(1)) if size_m else None

            idx_m = re.search(r"/Index\s*\[([^\]]+)\]", block)
            indices = []
            if idx_m:
                nums = [int(x) for x in idx_m.group(1).split()]
                indices = [(nums[i], nums[i+1]) for i in range(0, len(nums), 2)]
            elif size is not None:
                indices = [(0, size)]
            if indices:
                revision_indices.append(indices)

        objects_to_inspect = set()

        if revision_indices:

            last_rev_indices = max(revision_indices, key=lambda inds: max(start for (start,_count) in inds))

            for start, count in last_rev_indices:
                for obj_num in range(start, start+count):
                    objects_to_inspect.add(obj_num)
            objects_to_inspect = sorted(objects_to_inspect)
        else:
            objects_to_inspect = []
        
        object_details = []

        for obj_num in objects_to_inspect:
            try:
                out = subprocess.run(
                    ["python3", f"{TOOLS_PATH}/pdf-parser.py", "-o", str(obj_num), "-f", path],
                    capture_output=True,
                    text=True,
                    check=True,
                ).stdout
            except Exception as e:
                object_details.append({
                    "object_number": obj_num,
                    "object_type": None,
                    "object_subtype": None,
                    "object_text" : f"Erreur pdf-parser: {e}",
                })
                continue

            type_m = re.search(r"/Type\s*/([A-Za-z0-9]+)", out)
            subtype_m = re.search(r"/Subtype\s*/([A-Za-z0-9]+)", out)

            obj_type = type_m.group(1) if type_m else None 
            obj_subtype = subtype_m.group(1) if subtype_m else None

            texts = re.findall(r"\(([^)]*)\)\s*Tj?", out)
            if texts:

                text_decoded = texts[0] 
            
            else: text_decoded = None

            object_details.append({
                "object_number": obj_num,
                "object_type" : obj_type,
                "object_subtype" : obj_subtype,
                "object_text" : text_decoded
            })

        try:
            full_dump = subprocess.run([
                "python3", f"{TOOLS_PATH}/pdf-parser.py", path
            ],
            capture_output=True,
            text=True,
            check=True,
            ).stdout
        except Exception as e:
            full_dump = f"Erreur pdf-parser: {e}"
        
        return {
            "filename" : filename,
            "full_path" : path,
            "author" : author,
            "creator_tool" : creator,
            "producer" : producer,
            "creation_date" : creation_date,
            "mod_date" : mod_date,
            "startxrefs" : startxrefs,
            "startxrefs_count" : startxrefs_count,
            "trailers" : trailers,
            "trailers_count" : trailers_count,
            "revision_indices" : [
                [start, count] for rev in revision_indices for (start, count) in rev
            ],
            "objects_details" : object_details,
        }

    @task
    def build_excel(pdf_infos: list[dict], excel_path: str = OUTPUT_EXCEL) -> str | None :
        if not pdf_infos:
            return None
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            for info in pdf_infos:
                common = {
                    "filename" : info["filename"],
                    "full_path" : info["full_path"],
                    "author" : info["author"],
                    "creator_tool" : info["creator_tool"],
                    "producer" : info["producer"],
                    "creation_date" : info["creation_date"],
                    "mod_date" : info["mod_date"],
                    "startxrefs" : info["startxrefs"],
                    "startxrefs_count" : info["startxrefs_count"],
                    "trailers" : info["trailers"],
                    "trailers_count" : info["trailers_count"],
                    "revision_indices" : [f"[{start}, {count}]" for (start, count) in info["revision_indices"]],
                }

                rows = []

                for obj in info["objects_details"]:
                    row = {
                        **common,
                        "object_number" : obj["object_number"],
                        "object_type" : obj["object_type"],
                        "object_subtype" : obj["object_subtype"],
                        "object_text" : obj["object_text"],
                    }

                    rows.append(row)
                
                if not rows:
                    rows.append({
                        **common,
                        "object_number" : None,
                        "object_type" : None,
                        "object_subtype" : None,
                        "object_text" : None,
                    })
                
                df = pd.DataFrame(rows)
                sheet_name = os.path.splitext(info["filename"])[0][:21] or "PDF"
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return excel_path
    
    @task
    def timeline_excel(excel_path: str = OUTPUT_EXCEL) -> str :
        if not excel_path or not os.path.exists(excel_path):
            return None
        wb = load_workbook(excel_path)
        sheet_names = wb.sheetnames

        timeline_rows : list[dict] = []

        target_subtypes = {
            "sig", "signature", "text", "tx", "initial", "initials", "image"
        }
        for sheet_name in sheet_names:
            if sheet_name == "Timeline":
                continue
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            if df.empty:
                continue
            doc_name = df.get("filename", pd.Series([None])).iloc[0]
            creation_date = df.get("creation_date", pd.Series([None])).iloc[0]
            modif_date = df.get("mod_date", pd.Series([None])).iloc[0]

            if pd.notna(creation_date):
                timeline_rows.append({
                    "Date" : creation_date,
                    "Nom du document" : doc_name,
                    "Opération" : "Création",
                    "Valeur" : "",
                    "Notes supplémentaires" : "",
                })
            mod_effective = modif_date if pd.notna(modif_date) else creation_date

            for _, row in df.iterrows():
                subtype = str(row.get("object_subtype") or "").lower().strip()
                if subtype in target_subtypes and pd.notna(mod_effective):
                    obj_num = row.get("object_number")
                    value = row.get("object_text") if subtype in {"text", "tx"} else ""
                    timeline_rows.append({
                        "Date" : mod_effective,
                        "Nom du document" : doc_name,
                        "Opération" : "Modification",
                        "Élément" : subtype,
                        "Valeur" : value,
                        "Notes supplémentaires" : f"Objet n°{obj_num}"
                    })
        if not timeline_rows:
            return excel_path
        

        tl_df = pd.DataFrame(timeline_rows)
        if "Date" in tl_df.columns:
            try:
                tl_df["Date_sort"] = pd.to_datetime(tl_df["Date"], errors="coerce")
                tl_df = tl_df.sort_values("Date_sort").drop(columns=["Date_sort"])
            except Exception :
                pass

        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer : 
            tl_df.to_excel(writer, sheet_name="Timeline", index=False)

        return excel_path
    
    pdf_paths = list_pdfs()

    analyses = analyze_pdf.expand(path=pdf_paths)

    excel_file = build_excel(analyses)

    timeline = timeline_excel(excel_file)

dag = pdf_investigation()
